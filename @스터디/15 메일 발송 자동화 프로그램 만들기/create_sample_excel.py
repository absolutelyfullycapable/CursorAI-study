"""실습용 엑셀(메일실습용.xlsx) 생성 스크립트"""

from __future__ import annotations

import random
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

OUT = Path(__file__).resolve().parent / "메일실습용.xlsx"
PRACTICE_EMAIL = "fearnonemyself@gmail.com"

PRODUCTS = [
    "무선 이어폰",
    "노트북 거치대",
    "USB-C 케이블",
    "블루투스 스피커",
    "스마트워치 밴드",
    "휴대용 보조배터리",
    "기계식 키보드",
    "게이밍 마우스",
    "모니터 암",
    "웹캠",
    "HDMI 케이블",
    "노트북 파우치",
    "무선 충전기",
    "탁상 스탠드",
    "공기청정기 필터",
    "텀블러",
    "운동화",
    "백팩",
    "선풍기",
    "전기포트",
]
FIRST_NAMES = [
    "민수",
    "서연",
    "지훈",
    "하은",
    "도윤",
    "수아",
    "예준",
    "지우",
    "시우",
    "채원",
    "준서",
    "윤서",
    "현우",
    "소율",
    "건우",
    "지안",
    "우진",
    "다은",
    "선우",
    "예린",
    "태윤",
    "하린",
    "민준",
    "서진",
    "주원",
    "유진",
    "승현",
    "아린",
    "재민",
    "나윤",
]
LAST_NAMES = ["김", "이", "박", "최", "정", "강", "조", "윤", "장", "임", "한", "오", "서", "신", "권"]


def main() -> None:
    random.seed(42)
    wb = Workbook()
    ws = wb.active
    ws.title = "고객발송목록"

    headers = [
        "고객번호",
        "고객명",
        "이메일",
        "주문상품",
        "수량",
        "발송일",
        "운송장번호",
        "메일제목",
        "메일내용",
    ]
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )
    center = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin

    base_date = datetime(2026, 7, 1)
    used_names: set[str] = set()

    for i in range(1, 51):
        cust_no = f"C{i:04d}"
        while True:
            name = random.choice(LAST_NAMES) + random.choice(FIRST_NAMES)
            if name not in used_names:
                used_names.add(name)
                break
        product = random.choice(PRODUCTS)
        qty = random.randint(1, 5)
        ship_date = base_date + timedelta(days=random.randint(0, 13))
        tracking = f"5{random.randint(100000000000, 999999999999)}"
        subject = f"[배송안내] {name}님, 주문하신 {product}이(가) 발송되었습니다"
        body = (
            f"안녕하세요, {name}님.\n\n"
            f"주문하신 상품이 발송되어 안내드립니다.\n\n"
            f"- 고객번호: {cust_no}\n"
            f"- 주문상품: {product}\n"
            f"- 수량: {qty}개\n"
            f"- 발송일: {ship_date.strftime('%Y-%m-%d')}\n"
            f"- 운송장번호: {tracking}\n\n"
            f"배송 조회는 운송장번호로 가능합니다.\n"
            f"이용해주셔서 감사합니다."
        )
        row = [
            cust_no,
            name,
            PRACTICE_EMAIL,
            product,
            qty,
            ship_date.strftime("%Y-%m-%d"),
            tracking,
            subject,
            body,
        ]
        for col, val in enumerate(row, 1):
            cell = ws.cell(i + 1, col, val)
            cell.border = thin
            cell.alignment = center if col <= 7 else left_align

    widths = {
        "A": 12,
        "B": 10,
        "C": 28,
        "D": 18,
        "E": 8,
        "F": 12,
        "G": 18,
        "H": 50,
        "I": 55,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:I51"

    wb.save(OUT)
    print(f"생성 완료: {OUT}")


if __name__ == "__main__":
    main()
