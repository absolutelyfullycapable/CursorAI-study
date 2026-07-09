#!/usr/bin/env python3
"""금 시세 엑셀 데이터와 통계값을 이용해 시각화 PNG 이미지를 생성합니다."""

from __future__ import annotations

import argparse
import statistics
from datetime import datetime
from pathlib import Path

import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
from matplotlib.ticker import FuncFormatter
from openpyxl import load_workbook

DEFAULT_INPUT = Path(__file__).resolve().parent / "gold_prices_soongumnara.xlsx"
DEFAULT_OUTPUT_DIR = Path(__file__).resolve().parent / "charts"

DATA_SHEET = "금시세"

SERIES = [
    ("buy_pure", "내가 살 때 - 순금", "#C9A227", 2),
    ("sell_pure", "내가 팔 때 - 순금", "#E07B39", 3),
    ("sell_18k", "내가 팔 때 - 18K", "#5B8DEF", 4),
    ("sell_14k", "내가 팔 때 - 14K", "#6BA368", 5),
]

SHORT_LABELS = ["살 때\n순금", "팔 때\n순금", "팔 때\n18K", "팔 때\n14K"]
COLORS = [color for _, _, color, _ in SERIES]


def setup_korean_font() -> None:
    """macOS/Windows에서 한글이 깨지지 않도록 폰트를 설정합니다."""
    candidates = [
        "AppleGothic",
        "Malgun Gothic",
        "NanumGothic",
        "Noto Sans CJK KR",
        "Arial Unicode MS",
    ]
    available = {font.name for font in fm.fontManager.ttflist}
    for name in candidates:
        if name in available:
            plt.rcParams["font.family"] = name
            break
    plt.rcParams["axes.unicode_minus"] = False


def won_formatter(value: float, _: int) -> str:
    return f"{value:,.0f}"


def load_data(excel_path: Path) -> list[dict]:
    workbook = load_workbook(excel_path, data_only=True)
    if DATA_SHEET not in workbook.sheetnames:
        raise ValueError(f"'{DATA_SHEET}' 시트가 없습니다.")

    worksheet = workbook[DATA_SHEET]
    rows: list[dict] = []

    for row_index in range(2, worksheet.max_row + 1):
        date_text = worksheet.cell(row_index, 1).value
        values = [worksheet.cell(row_index, col).value for _, _, _, col in SERIES]
        if date_text is None or not all(isinstance(v, (int, float)) for v in values):
            continue

        parsed_date = datetime.strptime(str(date_text), "%Y.%m.%d")
        rows.append(
            {
                "date": parsed_date,
                "buy_pure": values[0],
                "sell_pure": values[1],
                "sell_18k": values[2],
                "sell_14k": values[3],
                "spread": values[0] - values[1],
            }
        )

    rows.sort(key=lambda row: row["date"])
    return rows


def calc_stats(rows: list[dict]) -> list[dict]:
    stats = []
    for key, label, color, _ in SERIES:
        values = [row[key] for row in rows]
        stats.append(
            {
                "key": key,
                "label": label,
                "color": color,
                "mean": statistics.mean(values),
                "median": statistics.median(values),
                "min": min(values),
                "max": max(values),
                "std": statistics.stdev(values) if len(values) > 1 else 0,
            }
        )
    return stats


def save_figure(fig: plt.Figure, output_dir: Path, filename: str) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / filename
    fig.savefig(path, dpi=150, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    return path


def chart_price_trend(rows: list[dict], output_dir: Path) -> Path:
    """1. 시계열 추이: 4개 금 시세 변화"""
    fig, ax = plt.subplots(figsize=(12, 6))
    for key, label, color, _ in SERIES:
        ax.plot(
            [row["date"] for row in rows],
            [row[key] for row in rows],
            label=label,
            color=color,
            linewidth=1.8,
            alpha=0.9,
        )

    ax.set_title("금 시세 일별 추이 (3.75g 기준)", fontsize=15, fontweight="bold", pad=12)
    ax.set_xlabel("고시 날짜")
    ax.set_ylabel("가격 (원)")
    ax.yaxis.set_major_formatter(FuncFormatter(won_formatter))
    ax.grid(True, linestyle="--", alpha=0.35)
    ax.legend(loc="upper right")
    fig.autofmt_xdate()
    return save_figure(fig, output_dir, "01_price_trend.png")


def chart_buy_vs_sell(rows: list[dict], output_dir: Path) -> Path:
    """2. 순금 매수 vs 매도 비교"""
    fig, ax = plt.subplots(figsize=(12, 6))
    dates = [row["date"] for row in rows]
    buy = [row["buy_pure"] for row in rows]
    sell = [row["sell_pure"] for row in rows]

    ax.plot(dates, buy, label="내가 살 때 - 순금", color="#C9A227", linewidth=2.2)
    ax.plot(dates, sell, label="내가 팔 때 - 순금", color="#E07B39", linewidth=2.2)
    ax.fill_between(dates, sell, buy, color="#F4E3A1", alpha=0.45, label="스프레드 구간")

    ax.set_title("순금 매수·매도 가격 비교", fontsize=15, fontweight="bold", pad=12)
    ax.set_xlabel("고시 날짜")
    ax.set_ylabel("가격 (원)")
    ax.yaxis.set_major_formatter(FuncFormatter(won_formatter))
    ax.grid(True, linestyle="--", alpha=0.35)
    ax.legend(loc="upper right")
    fig.autofmt_xdate()
    return save_figure(fig, output_dir, "02_buy_vs_sell.png")


def chart_spread_trend(rows: list[dict], output_dir: Path) -> Path:
    """3. 매수-매도 스프레드 추이"""
    spreads = [row["spread"] for row in rows]
    mean_spread = statistics.mean(spreads)

    fig, ax = plt.subplots(figsize=(12, 5))
    ax.plot(
        [row["date"] for row in rows],
        spreads,
        color="#8B5E3C",
        linewidth=2,
        marker="o",
        markersize=3,
    )
    ax.axhline(mean_spread, color="#D9534F", linestyle="--", linewidth=1.5, label=f"평균 {mean_spread:,.0f}원")

    ax.set_title("순금 매수-매도 스프레드 추이", fontsize=15, fontweight="bold", pad=12)
    ax.set_xlabel("고시 날짜")
    ax.set_ylabel("스프레드 (원)")
    ax.yaxis.set_major_formatter(FuncFormatter(won_formatter))
    ax.grid(True, linestyle="--", alpha=0.35)
    ax.legend(loc="upper right")
    fig.autofmt_xdate()
    return save_figure(fig, output_dir, "03_spread_trend.png")


def chart_mean_comparison(stats: list[dict], output_dir: Path) -> Path:
    """4. 항목별 평균 가격 막대 그래프"""
    fig, ax = plt.subplots(figsize=(10, 6))
    means = [item["mean"] for item in stats]
    bars = ax.bar(SHORT_LABELS, means, color=COLORS, edgecolor="white", linewidth=1.2)

    for bar, value in zip(bars, means):
        ax.text(
            bar.get_x() + bar.get_width() / 2,
            bar.get_height() + 3000,
            f"{value:,.0f}",
            ha="center",
            va="bottom",
            fontsize=10,
            fontweight="bold",
        )

    ax.set_title("항목별 평균 금 시세", fontsize=15, fontweight="bold", pad=12)
    ax.set_ylabel("평균 가격 (원)")
    ax.yaxis.set_major_formatter(FuncFormatter(won_formatter))
    ax.grid(axis="y", linestyle="--", alpha=0.35)
    ax.set_ylim(0, max(means) * 1.12)
    return save_figure(fig, output_dir, "04_mean_comparison.png")


def chart_min_max_range(stats: list[dict], output_dir: Path) -> Path:
    """5. 항목별 최소·평균·최대 비교"""
    x = range(len(stats))
    width = 0.25

    fig, ax = plt.subplots(figsize=(11, 6))
    mins = [item["min"] for item in stats]
    means = [item["mean"] for item in stats]
    maxs = [item["max"] for item in stats]

    ax.bar([i - width for i in x], mins, width=width, label="최소값", color="#7FB3D5")
    ax.bar(x, means, width=width, label="평균", color="#F5B041")
    ax.bar([i + width for i in x], maxs, width=width, label="최대값", color="#E74C3C")

    ax.set_title("항목별 최소·평균·최대 금 시세", fontsize=15, fontweight="bold", pad=12)
    ax.set_xticks(list(x))
    ax.set_xticklabels(SHORT_LABELS)
    ax.set_ylabel("가격 (원)")
    ax.yaxis.set_major_formatter(FuncFormatter(won_formatter))
    ax.grid(axis="y", linestyle="--", alpha=0.35)
    ax.legend()
    return save_figure(fig, output_dir, "05_min_mean_max.png")


def chart_boxplot(rows: list[dict], output_dir: Path) -> Path:
    """6. 항목별 가격 분포 (박스플롯)"""
    data = [[row[key] for row in rows] for key, _, _, _ in SERIES]
    labels = [label.replace("내가 ", "") for _, label, _, _ in SERIES]

    fig, ax = plt.subplots(figsize=(10, 6))
    box = ax.boxplot(
        data,
        tick_labels=labels,
        patch_artist=True,
        medianprops={"color": "#2C3E50", "linewidth": 2},
        whiskerprops={"linewidth": 1.2},
        capprops={"linewidth": 1.2},
    )

    for patch, color in zip(box["boxes"], COLORS):
        patch.set_facecolor(color)
        patch.set_alpha(0.65)

    ax.set_title("항목별 금 시세 분포", fontsize=15, fontweight="bold", pad=12)
    ax.set_ylabel("가격 (원)")
    ax.yaxis.set_major_formatter(FuncFormatter(won_formatter))
    ax.grid(axis="y", linestyle="--", alpha=0.35)
    return save_figure(fig, output_dir, "06_price_distribution.png")


def chart_std_comparison(stats: list[dict], output_dir: Path) -> Path:
    """7. 항목별 표준편차 (변동성) 비교"""
    fig, ax = plt.subplots(figsize=(10, 5))
    stds = [item["std"] for item in stats]
    bars = ax.barh(SHORT_LABELS, stds, color=COLORS, edgecolor="white")

    for bar, value in zip(bars, stds):
        ax.text(
            value + 150,
            bar.get_y() + bar.get_height() / 2,
            f"{value:,.1f}",
            va="center",
            fontsize=10,
        )

    ax.set_title("항목별 가격 변동성 (표준편차)", fontsize=15, fontweight="bold", pad=12)
    ax.set_xlabel("표준편차 (원)")
    ax.grid(axis="x", linestyle="--", alpha=0.35)
    return save_figure(fig, output_dir, "07_volatility.png")


def chart_daily_average(rows: list[dict], output_dir: Path) -> Path:
    """8. 일별 평균 순금 매수가"""
    daily: dict[str, list[float]] = {}
    for row in rows:
        key = row["date"].strftime("%Y.%m.%d")
        daily.setdefault(key, []).append(row["buy_pure"])

    dates = sorted(daily.keys())
    averages = [statistics.mean(daily[day]) for day in dates]
    parsed_dates = [datetime.strptime(day, "%Y.%m.%d") for day in dates]

    fig, ax = plt.subplots(figsize=(12, 5))
    ax.bar(parsed_dates, averages, width=0.8, color="#D4AF37", edgecolor="#B8962E")
    ax.plot(parsed_dates, averages, color="#8B6914", linewidth=1.5, marker="o", markersize=4)

    ax.set_title("일별 평균 순금 매수가", fontsize=15, fontweight="bold", pad=12)
    ax.set_xlabel("고시 날짜")
    ax.set_ylabel("일평균 가격 (원)")
    ax.yaxis.set_major_formatter(FuncFormatter(won_formatter))
    ax.grid(axis="y", linestyle="--", alpha=0.35)
    fig.autofmt_xdate()
    return save_figure(fig, output_dir, "08_daily_average_buy.png")


def chart_summary_dashboard(rows: list[dict], stats: list[dict], output_dir: Path) -> Path:
    """9. 통계 요약 대시보드"""
    fig = plt.figure(figsize=(14, 8))
    gs = fig.add_gridspec(2, 2, hspace=0.35, wspace=0.25)

    ax1 = fig.add_subplot(gs[0, :])
    for key, label, color, _ in SERIES[:2]:
        ax1.plot([row["date"] for row in rows], [row[key] for row in rows], label=label, color=color, linewidth=2)
    ax1.set_title("순금 매수·매도 추이", fontweight="bold")
    ax1.yaxis.set_major_formatter(FuncFormatter(won_formatter))
    ax1.grid(True, linestyle="--", alpha=0.3)
    ax1.legend(loc="upper right")
    fig.autofmt_xdate()

    ax2 = fig.add_subplot(gs[1, 0])
    ax2.bar(SHORT_LABELS, [item["mean"] for item in stats], color=COLORS)
    ax2.set_title("평균 시세 비교", fontweight="bold")
    ax2.yaxis.set_major_formatter(FuncFormatter(won_formatter))
    ax2.grid(axis="y", linestyle="--", alpha=0.3)

    ax3 = fig.add_subplot(gs[1, 1])
    spreads = [row["spread"] for row in rows]
    ax3.hist(spreads, bins=12, color="#E8C468", edgecolor="#B8962E")
    ax3.axvline(statistics.mean(spreads), color="#D9534F", linestyle="--", linewidth=2, label="평균")
    ax3.set_title("스프레드 분포", fontweight="bold")
    ax3.set_xlabel("스프레드 (원)")
    ax3.legend()

    period_start = rows[0]["date"].strftime("%Y.%m.%d")
    period_end = rows[-1]["date"].strftime("%Y.%m.%d")
    fig.suptitle(
        f"금 시세 통계 대시보드 ({period_start} ~ {period_end}, {len(rows)}건)",
        fontsize=16,
        fontweight="bold",
        y=0.98,
    )
    return save_figure(fig, output_dir, "09_summary_dashboard.png")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="금 시세 시각화 PNG 생성")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    setup_korean_font()

    rows = load_data(args.input)
    if not rows:
        raise SystemExit("시각화할 데이터가 없습니다.")

    stats = calc_stats(rows)
    charts = [
        chart_price_trend(rows, args.output_dir),
        chart_buy_vs_sell(rows, args.output_dir),
        chart_spread_trend(rows, args.output_dir),
        chart_mean_comparison(stats, args.output_dir),
        chart_min_max_range(stats, args.output_dir),
        chart_boxplot(rows, args.output_dir),
        chart_std_comparison(stats, args.output_dir),
        chart_daily_average(rows, args.output_dir),
        chart_summary_dashboard(rows, stats, args.output_dir),
    ]

    print(f"시각화 완료: {args.output_dir}")
    print(f"총 {len(charts)}개 PNG 파일 생성")
    for path in charts:
        print(f"  - {path.name}")


if __name__ == "__main__":
    main()
