#!/usr/bin/env python3
"""금 시세 엑셀 파일을 읽어 통계값을 계산하고 '통계' 시트에 기록합니다."""

from __future__ import annotations

import argparse
import statistics
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

DEFAULT_INPUT = Path(__file__).resolve().parent / "gold_prices_soongumnara.xlsx"
DATA_SHEET = "금시세"
STATS_SHEET = "통계"

PRICE_COLUMNS = [
    "내가 살 때(3.75g) - 순금",
    "내가 팔 때(3.75g) - 순금",
    "내가 팔 때(3.75g) - 18K",
    "내가 팔 때(3.75g) - 14K",
]


def load_price_data(worksheet) -> list[dict]:
    """'금시세' 시트에서 날짜와 가격 데이터를 읽습니다."""
    rows: list[dict] = []
    for row_index in range(2, worksheet.max_row + 1):
        values = [
            worksheet.cell(row_index, column).value
            for column in range(2, 2 + len(PRICE_COLUMNS))
        ]
        if not all(isinstance(value, (int, float)) for value in values):
            continue
        rows.append(
            {
                "date": worksheet.cell(row_index, 1).value,
                "values": values,
            }
        )
    return rows


def calc_column_stats(data_rows: list[dict], column_index: int) -> dict:
    """한 컬럼의 기본 통계값을 계산합니다."""
    values = [row["values"][column_index] for row in data_rows]
    minimum = min(values)
    maximum = max(values)

    min_row = next(row for row in data_rows if row["values"][column_index] == minimum)
    max_row = next(row for row in data_rows if row["values"][column_index] == maximum)

    return {
        "count": len(values),
        "mean": statistics.mean(values),
        "median": statistics.median(values),
        "min": minimum,
        "max": maximum,
        "range": maximum - minimum,
        "std": statistics.stdev(values) if len(values) > 1 else 0,
        "min_date": min_row["date"],
        "max_date": max_row["date"],
    }


def calc_spread_stats(data_rows: list[dict]) -> dict:
    """순금 매수가와 매도가 차이(스프레드) 통계를 계산합니다."""
    spreads = [
        row["values"][0] - row["values"][1]
        for row in data_rows
    ]
    return {
        "mean": statistics.mean(spreads),
        "median": statistics.median(spreads),
        "min": min(spreads),
        "max": max(spreads),
        "std": statistics.stdev(spreads) if len(spreads) > 1 else 0,
    }


def write_stats_sheet(workbook, data_rows: list[dict]) -> None:
    """통계 시트를 새로 만들거나 갱신합니다."""
    if STATS_SHEET in workbook.sheetnames:
        del workbook[STATS_SHEET]

    worksheet = workbook.create_sheet(STATS_SHEET, 1)
    column_stats = [calc_column_stats(data_rows, index) for index in range(len(PRICE_COLUMNS))]
    spread_stats = calc_spread_stats(data_rows)

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="D4AF37")
    label_fill = PatternFill("solid", fgColor="F5F0E1")

    worksheet["A1"] = "금 시세 통계 요약"
    worksheet["A1"].font = Font(bold=True, size=14)
    worksheet.merge_cells("A1:F1")

    worksheet["A2"] = (
        f"데이터 기간: {data_rows[-1]['date']} ~ {data_rows[0]['date']}"
    )
    worksheet["A3"] = f"총 고시 건수: {len(data_rows)}건"
    worksheet.merge_cells("A2:F2")
    worksheet.merge_cells("A3:F3")

    headers = ["통계 항목", *PRICE_COLUMNS]
    for column_index, header in enumerate(headers, 1):
        cell = worksheet.cell(5, column_index)
        cell.value = header
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    stat_rows = [
        ("건수", lambda stats: stats["count"], "0"),
        ("평균", lambda stats: round(stats["mean"]), "#,##0"),
        ("중앙값", lambda stats: round(stats["median"]), "#,##0"),
        ("최소값", lambda stats: stats["min"], "#,##0"),
        ("최대값", lambda stats: stats["max"], "#,##0"),
        ("범위(최대-최소)", lambda stats: stats["range"], "#,##0"),
        ("표준편차", lambda stats: round(stats["std"], 1), "#,##0.0"),
    ]

    current_row = 6
    for label, value_fn, number_format in stat_rows:
        label_cell = worksheet.cell(current_row, 1)
        label_cell.value = label
        label_cell.fill = label_fill
        label_cell.font = Font(bold=True)
        label_cell.alignment = Alignment(horizontal="center", vertical="center")
        label_cell.border = border

        for column_index, stats in enumerate(column_stats, 2):
            cell = worksheet.cell(current_row, column_index)
            cell.value = value_fn(stats)
            cell.number_format = number_format
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        current_row += 1

    detail_start = current_row + 1
    worksheet.cell(detail_start, 1).value = "최소값 고시일"
    worksheet.cell(detail_start, 1).font = Font(bold=True)
    worksheet.cell(detail_start + 1, 1).value = "최대값 고시일"
    worksheet.cell(detail_start + 1, 1).font = Font(bold=True)

    for column_index, stats in enumerate(column_stats, 2):
        min_cell = worksheet.cell(detail_start, column_index)
        max_cell = worksheet.cell(detail_start + 1, column_index)
        min_cell.value = stats["min_date"]
        max_cell.value = stats["max_date"]
        for cell in (min_cell, max_cell):
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

    spread_start = detail_start + 4
    worksheet.cell(spread_start, 1).value = "순금 매수-매도 스프레드 (살 때 - 팔 때)"
    worksheet.cell(spread_start, 1).font = Font(bold=True, size=12)
    worksheet.merge_cells(f"A{spread_start}:C{spread_start}")

    spread_rows = [
        ("평균 스프레드", round(spread_stats["mean"]), "#,##0"),
        ("중앙값 스프레드", round(spread_stats["median"]), "#,##0"),
        ("최소 스프레드", spread_stats["min"], "#,##0"),
        ("최대 스프레드", spread_stats["max"], "#,##0"),
        ("표준편차", round(spread_stats["std"], 1), "#,##0.0"),
    ]

    for offset, (label, value, number_format) in enumerate(spread_rows, 1):
        row_index = spread_start + offset
        label_cell = worksheet.cell(row_index, 1)
        value_cell = worksheet.cell(row_index, 2)
        unit_cell = worksheet.cell(row_index, 3)

        label_cell.value = label
        label_cell.fill = label_fill
        label_cell.font = Font(bold=True)
        label_cell.border = border

        value_cell.value = value
        value_cell.number_format = number_format
        value_cell.alignment = Alignment(horizontal="center")
        value_cell.border = border

        unit_cell.value = "원"
        unit_cell.border = border

    worksheet.column_dimensions["A"].width = 22
    for column in "BCDEF":
        worksheet.column_dimensions[column].width = 24


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="금 시세 엑셀 통계 계산")
    parser.add_argument(
        "--input",
        type=Path,
        default=DEFAULT_INPUT,
        help=f"입력 엑셀 파일 경로 (기본값: {DEFAULT_INPUT.name})",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if not args.input.exists():
        raise SystemExit(f"파일을 찾을 수 없습니다: {args.input}")

    workbook = load_workbook(args.input)
    if DATA_SHEET not in workbook.sheetnames:
        raise SystemExit(f"'{DATA_SHEET}' 시트가 없습니다.")

    data_rows = load_price_data(workbook[DATA_SHEET])
    if not data_rows:
        raise SystemExit("통계를 계산할 데이터가 없습니다.")

    write_stats_sheet(workbook, data_rows)
    workbook.save(args.input)

    print(f"통계 기록 완료: {args.input}")
    print(f"분석 건수: {len(data_rows)}건")
    for name, stats in zip(PRICE_COLUMNS, [calc_column_stats(data_rows, i) for i in range(4)]):
        print(
            f"- {name}: 평균 {round(stats['mean']):,}원 "
            f"(최소 {stats['min']:,} / 최대 {stats['max']:,})"
        )


if __name__ == "__main__":
    main()
