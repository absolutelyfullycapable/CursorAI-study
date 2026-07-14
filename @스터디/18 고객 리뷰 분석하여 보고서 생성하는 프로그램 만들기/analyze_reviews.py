#!/usr/bin/env python3
"""엑셀 고객 리뷰를 분석해 Claude API로 보고서를 생성합니다."""

from __future__ import annotations

import argparse
import os
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path
from statistics import mean

from anthropic import Anthropic
from dotenv import load_dotenv
from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent
DEFAULT_OUTPUT_DIR = BASE_DIR / "reports"
DEFAULT_EXCEL = BASE_DIR / "노트북 고객 리뷰 데이터.xlsx"

# 보고서 작성에 충분하면서 토큰 비용이 가장 낮은 모델
DEFAULT_MODEL = "claude-haiku-4-5"

REQUIRED_COLUMNS = ("리뷰내용", "평점")

SYSTEM_PROMPT = """당신은 경영진·임원 보고용 고객 리뷰 분석 컨설턴트입니다.
제공된 통계와 리뷰 원문만으로, 한눈에 읽히는 의사결정 지원 보고서를 마크다운으로 작성합니다.

작성 원칙:
- 숫자는 제공된 통계를 그대로 사용하고, 없는 사실을 지어내지 마세요.
- 인용은 실제 리뷰에서만 짧게 가져오세요.
- 가독성이 최우선입니다. 긴 문단보다 표·목록·짧은 불릿을 적극 사용하세요.
- 임원이 3분 안에 핵심을 파악할 수 있게 구성하세요.
- 과장·감정적 표현은 피하고, 근거 기반의 담담한 톤을 유지하세요."""

USER_PROMPT_TEMPLATE = """아래는 고객 리뷰 엑셀을 전처리한 결과입니다.
이 자료를 바탕으로 **경영진 보고용** 분석 보고서 1부를 작성해 주세요.

=== 기본 정보 ===
파일명: {filename}
분석 시각: {analyzed_at}
총 리뷰 수: {total_count}

=== 사전 계산된 핵심 지표 (마크다운) ===
{stats_markdown}

=== 평점 통계 (원문) ===
{rating_stats}

=== 기타 컬럼 요약 ===
{extra_stats}

=== 리뷰 원문 (평점 포함) ===
{reviews_text}

## 보고서 형식 (반드시 준수)

제목: `# 고객 리뷰 분석 보고서`

아래 섹션 순서·형식을 지키세요. **표와 목록을 충분히** 쓰세요.

### 1. 핵심 요약 (Executive Summary)
- 3~5개의 짧은 불릿으로만 작성
- 첫 불릿에 평균 평점·긍정/부정 비율을 포함
- 가장 시급한 이슈 1가지를 명시

### 2. 핵심 지표 대시보드
- 사전 계산된 핵심 지표 표를 **거의 그대로** 포함 (숫자 변경 금지)
- 이어서 평점 분포 표(| 평점 | 건수 | 비율 |)를 추가

### 3. 긍정 요인 TOP
- 표 형식: | 순위 | 요인 | 언급 요지 | 대표 근거(짧은 인용) |
- 4~6행

### 4. 부정 요인 · 리스크 TOP
- 표 형식: | 순위 | 이슈 | 영향도(상/중/하) | 대표 근거(짧은 인용) |
- 4~6행
- 영향도는 언급 빈도와 평점 하락 연관으로 판단

### 5. 고객 세그먼트 인사이트
- 연령대·성별 등 제공된 컬럼 기준으로 **표** 작성
  예: | 세그먼트 | 건수 | 특징 | 시사점 |
- 모델/제품군 인사이트가 있으면 별도 작은 표 또는 불릿 추가

### 6. 우선순위 액션 플랜
- 표 형식: | 우선순위 | 액션 | 기대 효과 | 난이도(상/중/하) |
- 실행 가능한 항목 4~5개
- 우선순위 1이 가장 시급

### 7. 대표 리뷰 인용
- 긍정 / 부정 각각 표로:
  | 구분 | 평점 | 인용 | 참고 정보(모델·연령대 등) |

### 8. 한 줄 결론
- 굵은 한 문장으로 의사결정 포인트 제시

분량: 1,500~2,500자 내외 (표·목록 포함).
서론·감사 인사·자기소개는 넣지 마세요.
"""


def find_header_row(ws, required: tuple[str, ...], max_scan: int = 10) -> int:
    """필수 컬럼명이 모두 있는 헤더 행 번호를 찾습니다 (1-based)."""
    for row_idx in range(1, min(max_scan, ws.max_row) + 1):
        values = [
            str(cell.value).strip() if cell.value is not None else ""
            for cell in ws[row_idx]
        ]
        if all(col in values for col in required):
            return row_idx
    raise ValueError(
        f"필수 컬럼 {list(required)}을(를) 헤더에서 찾지 못했습니다. "
        "엑셀 첫 행(또는 상단)에 '리뷰내용', '평점' 컬럼명이 있는지 확인해 주세요."
    )


def load_reviews(excel_path: Path) -> tuple[list[dict], list[str]]:
    """엑셀에서 리뷰 행과 컬럼 목록을 읽습니다. 컬럼은 이름으로 매칭합니다."""
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active
    header_row = find_header_row(ws, REQUIRED_COLUMNS)
    headers = [
        str(cell.value).strip() if cell.value is not None else f"열{i}"
        for i, cell in enumerate(ws[header_row], start=1)
    ]

    missing = [c for c in REQUIRED_COLUMNS if c not in headers]
    if missing:
        raise ValueError(f"필수 컬럼이 없습니다: {missing}")

    rows: list[dict] = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        record = {}
        for key, value in zip(headers, row):
            if key.startswith("열") and value is None:
                continue
            record[key] = value
        review = record.get("리뷰내용")
        rating = record.get("평점")
        if review is None or str(review).strip() == "":
            continue
        if rating is None or str(rating).strip() == "":
            continue
        try:
            record["평점"] = float(rating)
        except (TypeError, ValueError) as exc:
            raise ValueError(f"평점 값을 숫자로 변환할 수 없습니다: {rating!r}") from exc
        record["리뷰내용"] = str(review).strip()
        rows.append(record)

    if not rows:
        raise ValueError("유효한 리뷰 데이터가 없습니다. (리뷰내용·평점 모두 필요)")

    return rows, headers


def load_reviews_from_bytes(data: bytes, filename: str = "upload.xlsx") -> tuple[list[dict], list[str]]:
    """업로드된 엑셀 바이트에서 리뷰를 읽습니다."""
    from tempfile import NamedTemporaryFile

    suffix = Path(filename).suffix or ".xlsx"
    with NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
        tmp.write(data)
        tmp_path = Path(tmp.name)
    try:
        return load_reviews(tmp_path)
    finally:
        tmp_path.unlink(missing_ok=True)


def format_counter(counter: Counter, limit: int | None = None) -> str:
    items = counter.most_common(limit)
    if not items:
        return "- (데이터 없음)"
    return "\n".join(f"- {label}: {count}건" for label, count in items)


def compute_metrics(rows: list[dict], headers: list[str]) -> dict:
    """웹 UI·프롬프트용 핵심 지표를 계산합니다."""
    ratings = [r["평점"] for r in rows]
    total = len(ratings)
    positive = sum(1 for x in ratings if x >= 4)
    neutral = sum(1 for x in ratings if x == 3)
    negative = sum(1 for x in ratings if x <= 2)
    dist = Counter(ratings)

    segments: dict[str, list[dict]] = {}
    for col in ("연령대", "성별"):
        if col not in headers:
            continue
        counter = Counter(
            str(r.get(col)).strip()
            for r in rows
            if r.get(col) is not None and str(r.get(col)).strip()
        )
        segments[col] = [
            {"label": k, "count": v, "pct": round(v / total * 100, 1)}
            for k, v in counter.most_common()
        ]

    date_range = None
    if "구매일자" in headers:
        dates = []
        for r in rows:
            val = r.get("구매일자")
            if isinstance(val, datetime):
                dates.append(val.strftime("%Y-%m-%d"))
            elif val is not None and str(val).strip():
                dates.append(str(val).strip()[:10])
        if dates:
            date_range = {"from": min(dates), "to": max(dates)}

    return {
        "total": total,
        "avg_rating": round(mean(ratings), 2),
        "max_rating": max(ratings),
        "min_rating": min(ratings),
        "positive": positive,
        "neutral": neutral,
        "negative": negative,
        "positive_pct": round(positive / total * 100, 1),
        "neutral_pct": round(neutral / total * 100, 1),
        "negative_pct": round(negative / total * 100, 1),
        "rating_dist": [
            {
                "rating": float(k),
                "count": dist[k],
                "pct": round(dist[k] / total * 100, 1),
            }
            for k in sorted(dist)
        ],
        "segments": segments,
        "date_range": date_range,
        "columns": headers,
    }


def build_stats_markdown(metrics: dict) -> str:
    """임원 보고용 핵심 지표 표를 마크다운으로 만듭니다."""
    lines = [
        "| 지표 | 값 |",
        "|------|-----|",
        f"| 총 리뷰 수 | {metrics['total']}건 |",
        f"| 평균 평점 | **{metrics['avg_rating']}** / 5 |",
        f"| 최고 / 최저 | {metrics['max_rating']:.0f} / {metrics['min_rating']:.0f} |",
        f"| 긍정 (4~5점) | {metrics['positive']}건 ({metrics['positive_pct']}%) |",
        f"| 보통 (3점) | {metrics['neutral']}건 ({metrics['neutral_pct']}%) |",
        f"| 부정 (1~2점) | {metrics['negative']}건 ({metrics['negative_pct']}%) |",
    ]
    if metrics.get("date_range"):
        dr = metrics["date_range"]
        lines.append(f"| 구매 기간 | {dr['from']} ~ {dr['to']} |")
    return "\n".join(lines)


def build_stats(rows: list[dict], headers: list[str]) -> tuple[str, str]:
    """평점 통계와 기타 컬럼 요약을 텍스트로 만듭니다."""
    ratings = [r["평점"] for r in rows]
    rating_dist = Counter(ratings)
    positive = sum(1 for x in ratings if x >= 4)
    neutral = sum(1 for x in ratings if x == 3)
    negative = sum(1 for x in ratings if x <= 2)

    rating_stats = "\n".join(
        [
            f"평균 평점: {mean(ratings):.2f}",
            f"최고/최저: {max(ratings):.0f} / {min(ratings):.0f}",
            f"긍정(4~5점): {positive}건 ({positive / len(ratings) * 100:.1f}%)",
            f"보통(3점): {neutral}건 ({neutral / len(ratings) * 100:.1f}%)",
            f"부정(1~2점): {negative}건 ({negative / len(ratings) * 100:.1f}%)",
            "평점 분포:",
            format_counter(Counter({k: rating_dist[k] for k in sorted(rating_dist)})),
        ]
    )

    skip = set(REQUIRED_COLUMNS)
    extra_parts: list[str] = []
    for col in headers:
        if col in skip or col.startswith("열"):
            continue
        values = []
        for row in rows:
            val = row.get(col)
            if val is None or str(val).strip() == "":
                continue
            if isinstance(val, datetime):
                values.append(val.strftime("%Y-%m-%d"))
            else:
                values.append(str(val).strip())
        if not values:
            continue

        if all(_looks_like_date(v) for v in values[: min(5, len(values))]):
            dates = sorted(values)
            extra_parts.append(f"[{col}] 기간: {dates[0]} ~ {dates[-1]}")
            continue

        unique_ratio = len(set(values)) / len(values)
        counter = Counter(values)
        if unique_ratio > 0.8 and len(counter) > 15:
            extra_parts.append(
                f"[{col}] 고유값 {len(counter)}개 (거의 건마다 다름 — 개별 식별용으로 보임)"
            )
        else:
            extra_parts.append(f"[{col}] 분포:\n{format_counter(counter, limit=12)}")

    extra_stats = "\n\n".join(extra_parts) if extra_parts else "- (기타 컬럼 없음)"
    return rating_stats, extra_stats


def _looks_like_date(value: str) -> bool:
    if len(value) >= 8 and value[4:5] == "-" and value[7:8] == "-":
        return True
    return False


def format_reviews_for_prompt(rows: list[dict], headers: list[str]) -> str:
    """토큰을 아끼며 리뷰+보조 정보를 한 줄 요약으로 넣습니다."""
    meta_cols = [
        c
        for c in headers
        if c not in REQUIRED_COLUMNS and not c.startswith("열") and c != "고객명"
    ]
    prefer = ["노트북모델", "연령대", "성별", "구매일자", "고객ID"]
    use_cols = [c for c in prefer if c in meta_cols]
    for c in meta_cols:
        if c not in use_cols:
            use_cols.append(c)
    use_cols = use_cols[:5]

    lines: list[str] = []
    for i, row in enumerate(rows, start=1):
        meta = []
        for col in use_cols:
            val = row.get(col)
            if val is None or str(val).strip() == "":
                continue
            if isinstance(val, datetime):
                val = val.strftime("%Y-%m-%d")
            meta.append(f"{col}={val}")
        meta_str = ", ".join(meta)
        prefix = f"[{i}] 평점={row['평점']:.0f}"
        if meta_str:
            prefix += f" | {meta_str}"
        lines.append(f"{prefix}\n{row['리뷰내용']}")
    return "\n\n".join(lines)


def generate_report(
    client: Anthropic,
    model: str,
    filename: str,
    rows: list[dict],
    headers: list[str],
) -> str:
    metrics = compute_metrics(rows, headers)
    rating_stats, extra_stats = build_stats(rows, headers)
    reviews_text = format_reviews_for_prompt(rows, headers)
    user_content = USER_PROMPT_TEMPLATE.format(
        filename=filename,
        analyzed_at=datetime.now().strftime("%Y-%m-%d %H:%M"),
        total_count=len(rows),
        stats_markdown=build_stats_markdown(metrics),
        rating_stats=rating_stats,
        extra_stats=extra_stats,
        reviews_text=reviews_text,
    )
    response = client.messages.create(
        model=model,
        max_tokens=4096,
        system=SYSTEM_PROMPT,
        messages=[{"role": "user", "content": user_content}],
    )
    parts = []
    for block in response.content:
        if hasattr(block, "text"):
            parts.append(block.text)
    text = "\n".join(parts).strip()
    if not text:
        raise RuntimeError("Claude API가 빈 응답을 반환했습니다.")
    return text


def save_report(text: str, output_dir: Path, excel_stem: str) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = output_dir / f"리뷰분석보고서_{excel_stem}_{stamp}.md"
    path.write_text(text + "\n", encoding="utf-8")
    return path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="엑셀 고객 리뷰(리뷰내용·평점)를 분석해 Claude로 보고서를 생성합니다."
    )
    parser.add_argument(
        "excel",
        nargs="?",
        default=str(DEFAULT_EXCEL),
        help=f"분석할 엑셀 경로 (기본: {DEFAULT_EXCEL.name})",
    )
    parser.add_argument(
        "-o",
        "--output-dir",
        default=str(DEFAULT_OUTPUT_DIR),
        help="보고서 저장 폴더 (기본: reports/)",
    )
    parser.add_argument(
        "--model",
        default=DEFAULT_MODEL,
        help=f"Claude 모델 ID (기본값: {DEFAULT_MODEL})",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    load_dotenv(BASE_DIR / ".env")
    api_key = os.getenv("ANTHROPIC_API_KEY", "").strip()
    if not api_key:
        print(
            "오류: ANTHROPIC_API_KEY가 없습니다.\n"
            ".env 파일에 API 키를 넣어 주세요. (.env.example 참고)",
            file=sys.stderr,
        )
        return 1

    excel_path = Path(args.excel).expanduser().resolve()
    if not excel_path.is_file():
        print(f"오류: 엑셀 파일을 찾을 수 없습니다: {excel_path}", file=sys.stderr)
        return 1

    try:
        rows, headers = load_reviews(excel_path)
    except ValueError as exc:
        print(f"오류: {exc}", file=sys.stderr)
        return 1

    print(f"엑셀 로드: {excel_path.name} ({len(rows)}건)")
    print(f"컬럼: {', '.join(headers)}")
    print(f"모델: {args.model}")
    print("보고서 생성 중...")

    client = Anthropic(api_key=api_key)
    try:
        report = generate_report(
            client, args.model, excel_path.name, rows, headers
        )
    except Exception as exc:  # noqa: BLE001 — CLI에서 사용자 친화 메시지
        print(f"오류: Claude API 호출 실패 — {exc}", file=sys.stderr)
        return 1

    out_path = save_report(report, Path(args.output_dir), excel_path.stem)
    print("\n" + "=" * 60)
    print(report)
    print("=" * 60)
    print(f"\n저장 완료: {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
