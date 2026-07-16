"""
엑셀 고객 목록을 읽어 Gmail SMTP로 배송 안내 메일을 발송합니다.
실습용: 보내는/받는 주소 모두 동일 Gmail로 고정합니다.
"""

from __future__ import annotations

import os
import smtplib
import time
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent
EXCEL_PATH = BASE_DIR / "메일실습용.xlsx"
SHEET_NAME = "고객발송목록"

# 실습용: 발신/수신 모두 동일 주소
MAIL_ADDRESS = "fearnonemyself@gmail.com"

SMTP_HOST = "smtp.gmail.com"
SMTP_PORT = 587

# 스팸 오인 방지: 테스트 시 최대 발송 건수
TEST_LIMIT = 10
# 메일 사이 대기(초)
SEND_INTERVAL_SEC = 1.5


def load_customers(limit: int) -> list[dict]:
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(f"엑셀 파일을 찾을 수 없습니다: {EXCEL_PATH}")

    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
    headers = [cell.value for cell in ws[1]]

    required = ["고객번호", "고객명", "메일제목", "메일내용"]
    for col in required:
        if col not in headers:
            raise ValueError(f"엑셀에 '{col}' 컬럼이 없습니다. 현재 컬럼: {headers}")

    customers: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(v is None for v in row):
            continue
        record = dict(zip(headers, row))
        if not record.get("메일제목") or not record.get("메일내용"):
            continue
        customers.append(record)
        if len(customers) >= limit:
            break

    return customers


def build_message(subject: str, body: str) -> MIMEMultipart:
    msg = MIMEMultipart()
    msg["From"] = MAIL_ADDRESS
    msg["To"] = MAIL_ADDRESS
    msg["Subject"] = str(subject)
    msg.attach(MIMEText(str(body), "plain", "utf-8"))
    return msg


def send_emails(customers: list[dict], app_password: str) -> tuple[int, int]:
    success = 0
    failed = 0

    with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=30) as server:
        server.ehlo()
        server.starttls()
        server.ehlo()
        server.login(MAIL_ADDRESS, app_password)

        for i, customer in enumerate(customers, start=1):
            cust_no = customer.get("고객번호", "-")
            name = customer.get("고객명", "-")
            subject = customer["메일제목"]
            body = customer["메일내용"]

            try:
                msg = build_message(subject, body)
                server.sendmail(MAIL_ADDRESS, [MAIL_ADDRESS], msg.as_string())
                success += 1
                print(f"[{i}/{len(customers)}] 발송 성공 — {cust_no} {name}")
            except Exception as exc:  # noqa: BLE001 — 실습용: 건별 실패만 기록
                failed += 1
                print(f"[{i}/{len(customers)}] 발송 실패 — {cust_no} {name}: {exc}")

            if i < len(customers):
                time.sleep(SEND_INTERVAL_SEC)

    return success, failed


def main() -> None:
    load_dotenv(BASE_DIR / ".env")

    app_password = os.getenv("GMAIL_APP_PASSWORD", "").strip().replace(" ", "")
    if not app_password:
        raise SystemExit(
            "GMAIL_APP_PASSWORD가 없습니다.\n"
            "1) Google 계정 → 보안 → 2단계 인증 → 앱 비밀번호 생성\n"
            "2) 이 폴더의 .env.example을 참고해 .env에 GMAIL_APP_PASSWORD를 넣으세요."
        )

    print(f"엑셀: {EXCEL_PATH.name}")
    print(f"발신/수신: {MAIL_ADDRESS}")
    print(f"테스트 발송 한도: {TEST_LIMIT}건\n")

    customers = load_customers(TEST_LIMIT)
    if not customers:
        raise SystemExit("발송할 고객 데이터가 없습니다.")

    print(f"대상 {len(customers)}건 발송을 시작합니다...\n")
    success, failed = send_emails(customers, app_password)
    print(f"\n완료 — 성공 {success}건, 실패 {failed}건")


if __name__ == "__main__":
    main()
