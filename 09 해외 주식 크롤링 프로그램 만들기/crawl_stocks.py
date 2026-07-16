#!/usr/bin/env python3
"""Yahoo Finance 상승 주식 페이지에서 표 데이터를 크롤링해 엑셀 파일로 저장합니다."""

from __future__ import annotations

import argparse
import gzip
import re
import ssl
import urllib.error
import urllib.request
from pathlib import Path

from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

URL = "https://finance.yahoo.com/markets/stocks/gainers/"
TABLE_CLASS = "yf-1hgjbtd bd"
DEFAULT_OUTPUT = Path(__file__).resolve().parent / "stock_gainers_yahoo.xlsx"

COLUMNS = [
    "Symbol",
    "Name",
    "Price",
    "Change",
    "Change %",
    "Volume",
    "Avg Vol (3M)",
    "Market Cap",
    "P/E Ratio (TTM)",
    "52 Wk Change %",
    "52 Wk Range",
]

PLACEHOLDER_VALUES = {"--", "-", "N/A", "n/a"}


def fetch_html(url: str) -> str:
    """Yahoo Finance 페이지 HTML을 가져옵니다."""
    request = urllib.request.Request(
        url,
        headers={
            "User-Agent": (
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/120.0.0.0 Safari/537.36"
            ),
            "Accept": "text/html,application/xhtml+xml",
            "Accept-Language": "en-US,en;q=0.9",
            "Accept-Encoding": "gzip",
        },
    )

    contexts = [
        ssl.create_default_context(),
        ssl._create_unverified_context(),
    ]

    last_error: Exception | None = None
    for index, context in enumerate(contexts):
        try:
            with urllib.request.urlopen(request, context=context, timeout=30) as response:
                raw = response.read()
            if index == 1:
                print("주의: SSL 인증서 검증을 건너뛰고 페이지에 연결했습니다.")

            if raw[:2] == b"\x1f\x8b":
                return gzip.decompress(raw).decode("utf-8", errors="replace")
            return raw.decode("utf-8", errors="replace")
        except (urllib.error.URLError, OSError) as error:
            last_error = error

    raise RuntimeError("Yahoo Finance 페이지 호출에 실패했습니다.") from last_error


def clean_cell_text(text: str) -> str:
    """셀 텍스트에서 공백을 정리하고 의미 없는 값은 빈 문자열로 바꿉니다."""
    value = re.sub(r"\s+", " ", text).strip()
    if value in PLACEHOLDER_VALUES:
        return ""
    return value


def extract_price(value: str) -> str:
    """가격 열에서 첫 번째 숫자(현재가)만 추출합니다."""
    match = re.match(r"^[\d,]+(?:\.\d+)?", value.replace(",", ""))
    if not match:
        return clean_cell_text(value)
    return match.group(0)


def parse_gainers_table(html: str) -> list[dict[str, str]]:
    """yf-1hgjbtd bd 표에서 주식 데이터를 파싱합니다."""
    soup = BeautifulSoup(html, "html.parser")
    table = soup.find("table", class_=lambda value: value and TABLE_CLASS in value)
    if table is None:
        raise RuntimeError(f'class="{TABLE_CLASS}" 표를 찾을 수 없습니다.')

    rows: list[dict[str, str]] = []
    body_rows = table.find("tbody")
    if body_rows is None:
        return rows

    for row in body_rows.find_all("tr"):
        cells = row.find_all("td")
        if len(cells) < 12:
            continue

        values = [clean_cell_text(cell.get_text(" ", strip=True)) for cell in cells]

        # 두 번째 열은 차트 아이콘용 빈 칸이므로 건너뜁니다.
        record = {
            "Symbol": values[0],
            "Name": values[1],
            "Price": extract_price(values[3]),
            "Change": values[4],
            "Change %": values[5],
            "Volume": values[6],
            "Avg Vol (3M)": values[7],
            "Market Cap": values[8],
            "P/E Ratio (TTM)": values[9],
            "52 Wk Change %": values[10],
            "52 Wk Range": values[11],
        }

        if not record["Symbol"]:
            continue

        rows.append(record)

    return rows


def save_to_excel(rows: list[dict[str, str]], output_path: Path) -> None:
    """크롤링한 데이터를 엑셀 'Day Gainers' 시트에 저장합니다."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Day Gainers"
    worksheet.append(COLUMNS)

    header_fill = PatternFill("solid", fgColor="1A73E8")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for column in range(1, len(COLUMNS) + 1):
        cell = worksheet.cell(row=1, column=column)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for item in rows:
        worksheet.append([item.get(column, "") for column in COLUMNS])

    for row in worksheet.iter_rows(
        min_row=2,
        max_row=worksheet.max_row,
        min_col=1,
        max_col=len(COLUMNS),
    ):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

    worksheet.column_dimensions["A"].width = 10
    worksheet.column_dimensions["B"].width = 36
    for column in "CDEFGHIJK":
        worksheet.column_dimensions[column].width = 14

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Yahoo Finance 상승 주식 크롤링")
    parser.add_argument(
        "--url",
        default=URL,
        help=f"크롤링할 페이지 URL (기본값: {URL})",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help=f"저장할 엑셀 파일 경로 (기본값: {DEFAULT_OUTPUT.name})",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    print("Yahoo Finance 상승 주식 데이터를 가져오는 중...")
    html = fetch_html(args.url)
    rows = parse_gainers_table(html)
    if not rows:
        raise SystemExit("가져온 데이터가 없습니다.")

    save_to_excel(rows, args.output)

    print(f"저장 완료: {args.output}")
    print(f"총 {len(rows)}건 저장")


if __name__ == "__main__":
    main()
