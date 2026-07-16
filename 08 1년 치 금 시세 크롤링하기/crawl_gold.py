#!/usr/bin/env python3
"""순금나라 금 시세 API에서 데이터를 크롤링해 엑셀 파일로 저장합니다."""

from __future__ import annotations

import argparse
import json
import ssl
import urllib.error
import urllib.request
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

API_URL = "https://www.soongumnara.co.kr/api/price/chart/list"
DEFAULT_OUTPUT = Path(__file__).resolve().parent / "gold_prices_soongumnara.xlsx"

COLUMNS = [
    "고시날짜",
    "내가 살 때(3.75g) - 순금",
    "내가 팔 때(3.75g) - 순금",
    "내가 팔 때(3.75g) - 18K",
    "내가 팔 때(3.75g) - 14K",
]

FIELD_KEYS = ["s_pure", "p_pure", "p_18k", "p_14k"]


def fetch_gold_prices(
    data_date_start: str,
    data_date_end: str,
    search_term: str = "5M",
) -> list[dict]:
    """순금나라 API에서 금 시세 목록을 가져옵니다."""
    payload = json.dumps(
        {
            "srchDt": search_term,
            "type": "Au",
            "dataDateStart": data_date_start,
            "dataDateEnd": data_date_end,
        }
    ).encode("utf-8")

    request = urllib.request.Request(
        API_URL,
        data=payload,
        headers={
            "Content-Type": "application/json; charset=utf-8",
            "User-Agent": "Mozilla/5.0",
        },
        method="POST",
    )

    contexts = [
        ssl.create_default_context(),
        ssl._create_unverified_context(),
    ]

    last_error: Exception | None = None
    for index, context in enumerate(contexts):
        try:
            with urllib.request.urlopen(request, context=context, timeout=30) as response:
                body = json.loads(response.read().decode("utf-8"))
            if index == 1:
                print("주의: SSL 인증서 검증을 건너뛰고 API에 연결했습니다.")
            return body.get("list", [])
        except (urllib.error.URLError, json.JSONDecodeError) as error:
            last_error = error

    raise RuntimeError("금 시세 API 호출에 실패했습니다.") from last_error


def format_date(raw_date: str) -> str:
    """API 날짜(2026-07-09 15:32:15)를 표시 형식(2026.07.09)으로 바꿉니다."""
    if not raw_date:
        return ""
    return raw_date[:10].replace("-", ".")


def save_to_excel(rows: list[dict], output_path: Path) -> None:
    """크롤링한 데이터를 엑셀 '금시세' 시트에 저장합니다."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "금시세"
    worksheet.append(COLUMNS)

    header_fill = PatternFill("solid", fgColor="D4AF37")
    header_font = Font(bold=True)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for column in range(1, len(COLUMNS) + 1):
        cell = worksheet.cell(row=1, column=column)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for item in rows:
        worksheet.append(
            [format_date(item.get("date", ""))]
            + [item.get(key) for key in FIELD_KEYS]
        )

    for row in worksheet.iter_rows(
        min_row=2,
        max_row=worksheet.max_row,
        min_col=1,
        max_col=len(COLUMNS),
    ):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            if cell.column > 1 and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"

    worksheet.column_dimensions["A"].width = 14
    for column in "BCDE":
        worksheet.column_dimensions[column].width = 22

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def parse_args() -> argparse.Namespace:
    today = date.today()
    default_start = (today - timedelta(days=150)).strftime("%Y.%m.%d")
    default_end = today.strftime("%Y.%m.%d")

    parser = argparse.ArgumentParser(description="순금나라 금 시세 크롤링")
    parser.add_argument(
        "--limit",
        type=int,
        default=100,
        help="저장할 최대 건수 (기본값: 100)",
    )
    parser.add_argument(
        "--start",
        default=default_start,
        help=f"조회 시작일 YYYY.MM.DD (기본값: {default_start})",
    )
    parser.add_argument(
        "--end",
        default=default_end,
        help=f"조회 종료일 YYYY.MM.DD (기본값: {default_end})",
    )
    parser.add_argument(
        "--term",
        default="5M",
        help="조회 기간 코드 (기본값: 5M)",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help=f"저장할 엑셀 파일 경로 (기본값: {DEFAULT_OUTPUT.name})",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    print("금 시세 데이터를 가져오는 중...")
    rows = fetch_gold_prices(args.start, args.end, args.term)
    if not rows:
        raise SystemExit("가져온 데이터가 없습니다.")

    limited_rows = rows[: args.limit]
    save_to_excel(limited_rows, args.output)

    print(f"저장 완료: {args.output}")
    print(f"총 {len(limited_rows)}건 저장 (API 응답 {len(rows)}건 중 상위 {len(limited_rows)}건)")


if __name__ == "__main__":
    main()
